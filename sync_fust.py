import os
import openpyxl
import requests

try:
    import truststore
    truststore.inject_into_ssl()
except ImportError:
    pass

# Configurações
PASTA = os.path.dirname(__file__)
CAMINHO_UBS = os.path.join(PASTA, 'RESULTADO_PROVISORIO_DADOS.xlsx')
CAMINHO_SOLICITACOES = os.path.join(PASTA, 'SOLICITACOES.xlsx')
API_URL = 'https://darkgoldenrod-pelican-495804.hostingersite.com/sync-fust.php'
API_KEY = 'painel_fust_2026_key'


def to_str(val):
    return (val or '').strip() if isinstance(val, str) else (str(val).strip() if val is not None else '')


def to_int(val):
    try:
        return int(str(val).strip())
    except (TypeError, ValueError):
        return 0


def carregar_ubs():
    """Uma linha por estabelecimento (CNES). Base de referência dos elegíveis —
    quem tem QNT. COM EMP. VENC. = 1 já tem empresa vencedora manifestando interesse."""
    print(f'Lendo {CAMINHO_UBS}...')
    wb = openpyxl.load_workbook(CAMINHO_UBS, read_only=True, data_only=True)
    ws = wb['RESULTADO_GERAL']

    registros = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        registros.append({
            'cnes':               to_int(row[3]),
            'ibge':                to_int(row[1]),
            'uf':                  to_str(row[7]),
            'municipio':           to_str(row[9]),
            'nome_unidade':        to_str(row[4]),
            'tipo_unidade':        to_str(row[10]),
            'empresa_vencedora':   to_str(row[22]),
            'com_emp_venc':        1 if row[23] == 1 else 0,
        })
    wb.close()
    print(f'{len(registros)} estabelecimentos lidos.')
    return registros


def carregar_solicitacoes():
    """Uma linha por município que já manifestou adesão (Fase I e/ou Fase II)."""
    print(f'Lendo {CAMINHO_SOLICITACOES}...')
    wb = openpyxl.load_workbook(CAMINHO_SOLICITACOES, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    registros = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] is None:
            continue
        data_solicitacao = row[5]
        if hasattr(data_solicitacao, 'strftime'):
            data_solicitacao = data_solicitacao.strftime('%Y-%m-%d')
        elif data_solicitacao and '/' in str(data_solicitacao):
            d, m, y = str(data_solicitacao).split('/')
            data_solicitacao = f'{y}-{m}-{d}'
        else:
            data_solicitacao = None

        registros.append({
            'ibge':                to_int(row[2]),
            'uf':                  to_str(row[0]),
            'municipio':           to_str(row[1]),
            'numero_solicitacao':  to_str(row[4]),
            'data_solicitacao':    data_solicitacao,
            'situacao':            to_str(row[6]),
        })
    wb.close()
    print(f'{len(registros)} solicitações lidas.')
    return registros


def enviar(action, records):
    print(f'Enviando {len(records)} registros para {action}...')
    resp = requests.post(
        f'{API_URL}?action={action}',
        json={'records': records},
        headers={'X-Api-Key': API_KEY},
        timeout=180
    )
    resp.raise_for_status()
    result = resp.json()
    print(f'{action}: {result}')


if __name__ == '__main__':
    enviar('sync_ubs', carregar_ubs())
    enviar('sync_solicitacoes', carregar_solicitacoes())
    print('Sincronização concluída!')
